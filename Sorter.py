import json
from openpyxl import Workbook

workbook = Workbook(filename="Subject_65278_ChartEvents.xlsx")
sheet = workbook.active

meas_ID = []
chartDate = []
meas_Value = []

for x in range(1, workbook.max_row+1):
    print(workbook.cell(row=x, column=1).value)

    